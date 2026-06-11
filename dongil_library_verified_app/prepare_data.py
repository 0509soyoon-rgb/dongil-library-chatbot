import json
import re
import sqlite3
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
XLSX_PATH = BASE_DIR / 'data' / 'books.xlsx'
DB_PATH = BASE_DIR / 'data' / 'library.db'

CAREER_PROFILES = {
    '의학/보건': ['의학','의사','간호','간호사','보건','의료','병원','질병','바이러스','면역','인체','생명','생물','약','약학','뇌','건강'],
    '공학/AI': ['인공지능','AI','로봇','코딩','프로그래밍','컴퓨터','알고리즘','데이터','공학','기계','전자','소프트웨어','미래기술'],
    '수학/통계': ['수학','통계','확률','미적분','기하','논리','수학자','숫자'],
    '교육/교사': ['교육','교사','학교','수업','공부','청소년','아이','배움','학습','진로'],
    '법/사회': ['법','정의','헌법','인권','사회','정치','민주','국가','시민','범죄','검찰','판사','변호사'],
    '경제/경영': ['경제','경영','기업','돈','투자','금융','마케팅','회계','창업','시장','자본'],
    '문학/창작': ['소설','시','문학','작가','글쓰기','에세이','희곡','동화','고전','문장'],
    '심리/상담': ['심리','마음','감정','상담','관계','우울','불안','행복','성격'],
    '예술/디자인': ['미술','예술','디자인','건축','음악','영화','사진','그림','패션','공간'],
    '환경/생태': ['환경','생태','기후','지구','동물','식물','에너지','바다','숲','자연'],
    '역사/문화': ['역사','세계사','한국사','문화','문명','전쟁','제국','유럽','중국','일본'],
}

CAREER_ALIASES = {
    '의사':'의학/보건', '간호사':'의학/보건', '간호학과':'의학/보건', '약사':'의학/보건', '의대':'의학/보건', '약대':'의학/보건', '보건':'의학/보건',
    'ai':'공학/AI', '인공지능':'공학/AI', '개발자':'공학/AI', '컴공':'공학/AI', '컴퓨터공학':'공학/AI', '공학':'공학/AI', '데이터':'공학/AI',
    '수학':'수학/통계', '수학교사':'수학/통계', '통계':'수학/통계', '수학교육':'수학/통계',
    '교사':'교육/교사', '교육학':'교육/교사', '초등교육':'교육/교사', '사범대':'교육/교사',
    '법':'법/사회', '법학':'법/사회', '변호사':'법/사회', '판사':'법/사회', '정치':'법/사회', '사회':'법/사회',
    '경영':'경제/경영', '경제':'경제/경영', '회계':'경제/경영', '마케팅':'경제/경영',
    '문예창작':'문학/창작', '국어교육':'문학/창작', '작가':'문학/창작', '문학':'문학/창작',
    '심리':'심리/상담', '상담':'심리/상담', '상담사':'심리/상담',
    '디자인':'예술/디자인', '건축':'예술/디자인', '미술':'예술/디자인', '음악':'예술/디자인',
    '환경':'환경/생태', '생태':'환경/생태', '기후':'환경/생태',
    '역사':'역사/문화', '문화':'역사/문화', '한국사':'역사/문화', '세계사':'역사/문화'
}

KDC = [
    ('000', '총류/정보'), ('100', '철학/심리'), ('200', '종교'), ('300', '사회과학'), ('400', '자연과학'),
    ('500', '기술과학'), ('600', '예술'), ('700', '언어'), ('800', '문학'), ('900', '역사/지리')
]

def clean(v):
    if v is None:
        return ''
    return str(v).strip()

def norm(s):
    return re.sub(r'[^0-9a-zA-Z가-힣]+', '', (s or '').lower())

def kdc_genre(callno, title):
    s = clean(callno)
    m = re.match(r'(\d{3})', s)
    if m:
        n = int(m.group(1))
        if 0 <= n < 100: return '총류/정보'
        if 100 <= n < 200: return '철학/심리'
        if 200 <= n < 300: return '종교'
        if 300 <= n < 400: return '사회과학'
        if 400 <= n < 500: return '자연과학'
        if 500 <= n < 600: return '기술과학'
        if 600 <= n < 700: return '예술'
        if 700 <= n < 800: return '언어'
        if 800 <= n < 900: return '문학'
        if 900 <= n < 1000: return '역사/지리'
    t = title.lower()
    if any(x in t for x in ['소설','문학','시집','에세이']): return '문학'
    if any(x in t for x in ['수학','과학','물리','화학','생명']): return '자연과학'
    return '미분류'

def infer_tags(title, author, publisher, genre):
    text = f'{title} {author} {publisher} {genre}'.lower()
    tags = []
    for career, kws in CAREER_PROFILES.items():
        if any(k.lower() in text for k in kws):
            tags.append(career)
    if genre and genre not in tags:
        tags.insert(0, genre)
    if any(k in text for k in ['청소년','고등학생','10대','진로']): tags.append('청소년')
    if any(k in text for k in ['고전','세계문학','한국문학']): tags.append('고전/문학')
    return list(dict.fromkeys(tags))[:8]

def infer_level(title, callno, genre, year):
    """강화된 난이도 분류.
    원본 책 정보를 바꾸지 않고 제목/청구기호/분야 단서만으로 보조 난이도를 산정합니다.
    쉬움/보통이 과도하게 많아지지 않도록 전문 분야와 고전/원전 계열은 상향합니다.
    """
    text = f'{title} {callno} {genre}'.lower()
    score = 3

    # 쉬운 접근 단서
    easy_words = ['그림책','동화','만화','웹툰','쉽게','쉬운','처음','입문','청소년','10대','어린이','초등','재미있는','한눈에','교과서 밖']
    if any(k in text for k in easy_words):
        score -= 1

    # 가벼운 장르 단서
    if genre == '문학':
        score -= 1
    if any(k in text for k in ['시집','에세이','수필','편지','일기']):
        score -= 1

    # 심화/전문 단서
    hard_words = ['개론','원론','철학','비판','이론','심화','논문','전문','고급','원전','고전','사상','담론','인식론','존재론','사회학','정치학','경제학','법학','물리학','화학','생물학','수학','통계','미적분','유전','뇌과학','양자','상대성','문명','세계사','한국사','자본','민주주의']
    if any(k in text for k in hard_words):
        score += 1
    very_hard_words = ['순수이성비판','자본론','국부론','종의 기원','군주론','논어','맹자','장자','도덕경','방법서설','꿈의 해석','코스모스','총균쇠','사피엔스','이기적 유전자']
    if any(k in text for k in very_hard_words):
        score += 1

    # KDC 계열로 난이도 보정
    if genre in ['철학/심리','사회과학','자연과학','기술과학','역사/지리']:
        score += 1
    if genre in ['총류/정보','언어','종교']:
        score += 0
    if genre == '예술':
        score -= 0

    # 제목이 아주 짧은 문학 작품은 쉬움으로만 단정하지 않음
    # 고전/세계문학/전집 계열은 상향
    if any(k in text for k in ['세계문학','한국문학','고전문학','문학전집','전집','선집']):
        score += 1

    # 입문 단서가 있어도 전문 단서가 많으면 최소 보통 이상
    if any(k in text for k in ['철학','경제학','물리학','화학','수학','통계','법학','정치학']) and score < 3:
        score = 3

    return max(1, min(5, score))

def difficulty_label(level):
    return ['','입문','쉬움','보통','도전','심화'][level]

def infer_careers(tags, genre, title):
    careers = []
    for t in tags:
        if t in CAREER_PROFILES:
            careers.append(t)
    if genre == '문학': careers += ['문학/창작','교육/교사']
    elif genre == '자연과학': careers += ['수학/통계','의학/보건','환경/생태']
    elif genre == '기술과학': careers += ['공학/AI','의학/보건','예술/디자인']
    elif genre == '사회과학': careers += ['법/사회','경제/경영','교육/교사']
    elif genre == '철학/심리': careers += ['심리/상담','교육/교사']
    elif genre == '예술': careers += ['예술/디자인']
    elif genre == '역사/지리': careers += ['역사/문화','법/사회']
    return list(dict.fromkeys(careers))[:5]

def intro(title, genre):
    """책 내용 1~2줄 소개.
    출판사/교보문고 전문 줄거리를 저장하지 않는 대신, 원본 자료명과 분류를 바탕으로 거짓 제목을 만들지 않는 안전한 자동 소개를 생성합니다.
    """
    if genre == '문학':
        return '인물과 사건을 따라가며 삶, 감정, 관계를 생각해 볼 수 있는 문학 분야 소장도서입니다. 작품의 분위기와 주제를 중심으로 읽기 좋습니다.'
    if genre == '자연과학':
        return '과학 개념과 자연 현상을 이해하는 데 도움이 되는 자연과학 분야 소장도서입니다. 핵심 개념을 정리하며 읽기 좋습니다.'
    if genre == '기술과학':
        return '기술, 의학, 공학, 생활 속 문제 해결을 다루는 기술과학 분야 소장도서입니다. 진로 탐색과 탐구활동에 연결하기 좋습니다.'
    if genre == '사회과학':
        return '사회 문제, 제도, 경제, 정치, 교육 등 현실 세계를 이해하는 데 도움이 되는 사회과학 분야 소장도서입니다.'
    if genre == '철학/심리':
        return '생각, 가치관, 마음, 관계를 깊게 바라보는 철학·심리 분야 소장도서입니다. 자기 이해와 토론 주제 만들기에 좋습니다.'
    if genre == '역사/지리':
        return '역사적 사건, 인물, 문화와 지역의 흐름을 살펴볼 수 있는 역사·지리 분야 소장도서입니다.'
    if genre == '예술':
        return '미술, 음악, 디자인, 건축 등 표현과 창작을 이해하는 데 도움이 되는 예술 분야 소장도서입니다.'
    if genre == '언어':
        return '언어 표현, 말과 글, 소통 방식을 이해하는 데 도움이 되는 언어 분야 소장도서입니다.'
    return '동일여고 원본 소장도서 목록에 있는 자료입니다. 제목과 분류를 바탕으로 탐색할 수 있습니다.'

def reason(title, genre, careers):
    c = ', '.join(careers[:2]) if careers else '진로 탐색'
    return f'이 책은 {genre} 분야의 배경지식을 넓히고, {c}와 연결해 생각해 보기 좋아 추천합니다.'

def topics(title, genre, careers):
    base = {
        '문학':['인물의 성장과 갈등 분석','작품 속 시대·사회 배경 탐구','문학 작품의 주제 의식 비교'],
        '자연과학':['핵심 과학 개념 정리와 실제 사례 탐구','과학기술이 사회에 미치는 영향','관련 실험·관찰 주제 설계'],
        '기술과학':['기술 발전이 생활을 바꾸는 방식','공학적 문제 해결 사례 분석','미래 기술의 장단점 탐구'],
        '사회과학':['현대 사회문제의 원인과 대안','정책과 시민의 역할 탐구','통계·기사 자료를 활용한 사회 분석'],
        '철학/심리':['인간의 선택과 가치관 탐구','감정과 관계를 설명하는 심리 개념','윤리적 딜레마 토론'],
        '역사/지리':['역사적 사건의 원인과 결과','문화권 비교 탐구','현재 사회와 역사적 사건의 연결'],
        '예술':['작품과 시대적 배경 연결','예술 표현 방식 비교','디자인과 생활 문제 해결'],
        '언어':['언어 표현 방식과 소통','말과 글의 사회적 역할','번역과 문화 차이 탐구'],
    }
    return base.get(genre, ['책의 핵심 주제 정리','진로와 연결되는 질문 만들기','비슷한 소장도서와 비교 읽기'])[:3]

def reading_type_tags(code):
    tags=[]
    if code[0]=='S': tags += ['문학/창작','교육/교사','심리/상담']
    else: tags += ['수학/통계','공학/AI','자연과학','경제/경영']
    if code[1]=='L': tags += ['청소년','문학']
    else: tags += ['철학/심리','자연과학','사회과학']
    if code[2]=='R': tags += ['법/사회','역사/문화','의학/보건']
    else: tags += ['문학/창작','공학/AI','예술/디자인']
    if code[3]=='G': tags += ['진로','교육/교사','공학/AI','의학/보건']
    else: tags += ['문학/창작','심리/상담','예술/디자인']
    return tags

def main():
    if not XLSX_PATH.exists():
        raise SystemExit(f'엑셀 파일이 없습니다: {XLSX_PATH}')
    if DB_PATH.exists(): DB_PATH.unlink()
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [clean(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h:i for i,h in enumerate(headers)}
    required = ['No','자료유형','등록번호','자료명','저자','출판사','출판년도','청구기호','등록일','수정일','자료상태','소장처']
    missing = [c for c in required if c not in idx]
    if missing:
        raise SystemExit('필수 열 누락: '+', '.join(missing))
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE books (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        original_no INTEGER,
        material_type TEXT,
        reg_no TEXT,
        title TEXT NOT NULL,
        title_norm TEXT NOT NULL,
        author TEXT,
        publisher TEXT,
        pub_year TEXT,
        call_no TEXT,
        registered_at TEXT,
        updated_at TEXT,
        status TEXT,
        location TEXT,
        genre TEXT,
        tags_json TEXT,
        difficulty INTEGER,
        difficulty_label TEXT,
        careers_json TEXT,
        intro TEXT,
        recommend_reason TEXT,
        topics_json TEXT,
        data_note TEXT
    )''')
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = clean(row[idx['자료명']])
        if not title:
            continue
        author = clean(row[idx['저자']])
        publisher = clean(row[idx['출판사']])
        call_no = clean(row[idx['청구기호']])
        pub_year = clean(row[idx['출판년도']])
        genre = kdc_genre(call_no, title)
        tags = infer_tags(title, author, publisher, genre)
        level = infer_level(title, call_no, genre, pub_year)
        careers = infer_careers(tags, genre, title)
        rows.append((
            row[idx['No']], clean(row[idx['자료유형']]), clean(row[idx['등록번호']]), title, norm(title),
            author, publisher, pub_year, call_no, clean(row[idx['등록일']]), clean(row[idx['수정일']]),
            clean(row[idx['자료상태']]), clean(row[idx['소장처']]), genre, json.dumps(tags, ensure_ascii=False), level,
            difficulty_label(level), json.dumps(careers, ensure_ascii=False), intro(title, genre), reason(title, genre, careers),
            json.dumps(topics(title, genre, careers), ensure_ascii=False), '원본 정보는 업로드된 동일여고 소장도서 엑셀 기준입니다. 장르·태그·난이도·추천 이유·탐구주제·간략 소개는 청구기호와 제목 기반 자동 보조 분류입니다.'
        ))
    c.executemany('''INSERT INTO books (
        original_no, material_type, reg_no, title, title_norm, author, publisher, pub_year, call_no,
        registered_at, updated_at, status, location, genre, tags_json, difficulty, difficulty_label,
        careers_json, intro, recommend_reason, topics_json, data_note
    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', rows)
    c.execute('CREATE INDEX idx_books_title_norm ON books(title_norm)')
    c.execute('CREATE INDEX idx_books_title ON books(title)')
    c.execute('CREATE INDEX idx_books_genre ON books(genre)')
    c.execute('CREATE INDEX idx_books_difficulty ON books(difficulty)')
    meta = {
        'created_at': datetime.now().isoformat(timespec='seconds'),
        'book_count': len(rows),
        'source_file': XLSX_PATH.name,
        'rules': ['원본 엑셀에 있는 도서만 추천', '제목/저자/출판사/청구기호 원본 유지', '보조 정보는 자동 분류로 표시']
    }
    c.execute('CREATE TABLE meta (key TEXT PRIMARY KEY, value TEXT)')
    for k,v in meta.items():
        c.execute('INSERT INTO meta VALUES (?,?)', (k, json.dumps(v, ensure_ascii=False)))
    conn.commit(); conn.close()
    print(f'완료: {len(rows)}권을 {DB_PATH}에 저장했습니다.')

if __name__ == '__main__':
    main()
